# AI Case Engine MVP Step 1 v7 - template master logic
# Datei: run_engine_mvp_v7.py
#
# Leg diese Datei direkt in deinen Hauptordner:
# C:\Users\Admin\Desktop\AI Case Engine
#
# Danach per Doppelklick auf run_engine_mvp.bat starten.
#
# Erwartete Ordnerstruktur:
# C:\Users\Admin\Desktop\AI Case Engine
# â”œâ”€â”€ 01_Input
# â”œâ”€â”€ 02_Regelmatrix
# â””â”€â”€ 03_Output

from __future__ import annotations

import re
import sys
from pathlib import Path
from io import BytesIO
from datetime import datetime
from typing import Any, Dict, List, Tuple

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule, DataBarRule, ColorScaleRule
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "01_Input"
RULE_DIR = BASE_DIR / "02_Regelmatrix"
OUTPUT_DIR = BASE_DIR / "03_Output"
OUTPUT_FILE = OUTPUT_DIR / "case_output_mvp.xlsx"
DEFAULT_RULEBOOK_FILE = RULE_DIR / "AI_Case_Engine_rulebook_MVP_Step1_Fleet_v7_TemplateMaster.xlsx"

REQUIRED_INPUT_COLUMNS = [
    "case_id",
    "customer_name",
    "vehicle_model",
    "vin",
    "case_subject",
    "case_text",
    "previous_cases",
    "language",
]

REQUIRED_RULEBOOK_SHEETS = [
    "business_value_matrix",
    "customer_type_rules",
    "fleet_value_rules",
    "urgency_keywords",
    "escalation_keywords",
    "emotion_keywords",
    "case_type_rules",
    "score_weights",
    "scoring_config",
    "priority_floor_rules",
    "template_master",
]

OPTIONAL_RULEBOOK_SHEETS = ["template_mapping"]

AGENT_VIEW_COLUMNS = [
    "case_id",
    "case_priority_class",
    "escalation_risk_level",
    "priority_score",
    "urgency_level",
    "business_value_level",
    "customer_name",
    "vehicle_model",
    "recommended_template_id",
    "tone_level",
    "recommended_next_action",
    "agent_warning",
    "deescalation_phrases",
    "forbidden_claims",
    "recommended_customer_reply",
    "template_selection_reason",
    "decision_reason",
]


# ------------------------------------------------------------
# Basis-Helfer
# ------------------------------------------------------------

def log(msg: str) -> None:
    print(msg, flush=True)


def normalize_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    text = str(value).lower()
    text = text.replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def to_float(value: Any, default: float = 0.0) -> float:
    try:
        if pd.isna(value):
            return default
        return float(str(value).replace(",", "."))
    except Exception:
        return default


def to_int(value: Any, default: int = 0) -> int:
    try:
        if pd.isna(value):
            return default
        return int(float(str(value).replace(",", ".")))
    except Exception:
        return default


def contains_match(text: str, trigger: str) -> bool:
    trigger_norm = normalize_text(trigger)
    if not trigger_norm:
        return False
    return trigger_norm in text


def find_first_file(folder: Path, patterns: List[str], preferred_keyword: str) -> Path:
    files: List[Path] = []
    for pattern in patterns:
        files.extend(folder.glob(pattern))

    files = [f for f in files if not f.name.startswith("~$")]

    if not files:
        raise FileNotFoundError(f"Keine passende Datei in {folder} gefunden.")

    files = sorted(
        files,
        key=lambda f: (0 if preferred_keyword.lower() in f.name.lower() else 1, f.name.lower())
    )
    return files[0]


def find_input_file() -> Path:
    return find_first_file(INPUT_DIR, ["*.xlsx", "*.xls", "*.csv"], "input")


def find_rulebook_file() -> Path:
    # v7 bevorzugt das Rulebook mit template_master.
    # Falls mehrere Rulebooks im Ordner liegen, wird zuerst ein TemplateMaster-Rulebook genommen.
    return find_first_file(RULE_DIR, ["*.xlsx", "*.xls"], "TemplateMaster")


def read_cases(path: Path | BytesIO | Any) -> pd.DataFrame:
    """Read MVP case input and validate the exact v7 columns."""
    suffix = str(getattr(path, "suffix", "") or getattr(path, "name", "")).lower()
    if suffix.endswith(".csv"):
        df = pd.read_csv(path, dtype=str).fillna("")
    else:
        xls = pd.ExcelFile(path)
        sheet = "case_input" if "case_input" in xls.sheet_names else xls.sheet_names[0]
        df = pd.read_excel(path, sheet_name=sheet, dtype=str).fillna("")

    missing = [c for c in REQUIRED_INPUT_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError("Folgende Input-Spalten fehlen: " + ", ".join(missing))

    return df[REQUIRED_INPUT_COLUMNS].copy()


def read_rulebook(path: Path | BytesIO | Any = DEFAULT_RULEBOOK_FILE) -> Dict[str, pd.DataFrame]:
    """Read the real MVP v7 rulebook. The Excel rulebook remains the source of truth."""
    rules: Dict[str, pd.DataFrame] = {}
    xls = pd.ExcelFile(path)

    missing = [sheet for sheet in REQUIRED_RULEBOOK_SHEETS if sheet not in xls.sheet_names]
    if missing:
        raise ValueError("Folgende Rulebook-Sheets fehlen: " + ", ".join(missing))

    for sheet in REQUIRED_RULEBOOK_SHEETS + OPTIONAL_RULEBOOK_SHEETS:
        if sheet in xls.sheet_names:
            rules[sheet] = pd.read_excel(path, sheet_name=sheet, dtype=str).fillna("")
            rules[sheet] = rules[sheet].loc[
                ~(rules[sheet].astype(str).apply(lambda r: "".join(r).strip(), axis=1) == "")
            ].copy()
        else:
            rules[sheet] = pd.DataFrame()

    return rules


def source_value(case: pd.Series, source_field: str) -> str:
    source_field = str(source_field).strip()
    if source_field in ["all_text", "text_bundle"]:
        return normalize_text(f"{case.get('case_subject', '')} {case.get('case_text', '')}")
    return normalize_text(case.get(source_field, ""))


def score_to_level(score: float, config: pd.DataFrame, score_type: str, fallback: str) -> str:
    if config.empty:
        return fallback

    if "score_type" not in config.columns:
        return fallback

    part = config[config["score_type"].str.lower() == score_type.lower()].copy()
    if part.empty:
        return fallback

    for _, row in part.iterrows():
        mn = to_float(row.get("min_score", 0))
        mx = to_float(row.get("max_score", 0))
        out = str(row.get("output_class", fallback))
        if mn <= score <= mx:
            return out

    return fallback


def priority_rank(priority_class: str) -> int:
    return {"P1": 1, "P2": 2, "P3": 3, "P4": 4}.get(str(priority_class).upper(), 9)


def risk_rank(risk_level: str) -> int:
    return {"red": 1, "orange": 2, "yellow": 3, "green": 4}.get(str(risk_level).lower(), 9)


def max_priority(p1: str, p2: str) -> str:
    """Gibt die hÃ¶here PrioritÃ¤t zurÃ¼ck: P1 ist hÃ¶her als P2."""
    return p1 if priority_rank(p1) <= priority_rank(p2) else p2


# ------------------------------------------------------------
# Fachlogik
# ------------------------------------------------------------

def derive_business_value(case: pd.Series, matrix: pd.DataFrame) -> Tuple[float, str, str]:
    if matrix.empty:
        return 0.0, "unknown", "Keine business_value_matrix vorhanden"

    model = normalize_text(case.get("vehicle_model", ""))
    df = matrix.copy()

    if "vehicle_model" not in df.columns:
        return 0.0, "unknown", "Spalte vehicle_model fehlt in business_value_matrix"

    df["_model"] = df["vehicle_model"].apply(normalize_text)

    match = df[df["_model"] == model]

    if match.empty and model:
        match = df[df["_model"].apply(lambda x: bool(x) and (x in model or model in x))]

    if match.empty:
        return 0.0, "unknown", "Kein Modell-Mapping gefunden"

    row = match.iloc[0]
    score = to_float(
        row.get("business_value_score", row.get("base_business_value_score", 0))
    )
    level = str(row.get("business_value_level", "unknown"))
    reason = str(row.get("reason", ""))
    return min(score, 10.0), level, reason


def apply_keyword_rules(
    case: pd.Series,
    rules: pd.DataFrame,
    score_col: str,
    level_col: str | None = None,
    max_score: float = 10.0
) -> Tuple[float, str, List[str]]:
    """
    MVP-Logik:
    - Es werden alle Treffer gesammelt.
    - FÃ¼r den Score zÃ¤hlt der hÃ¶chste Treffer, nicht die Summe.
      Dadurch bleibt die 0-10-Logik stabil.
    """
    if rules.empty:
        return 0.0, "none", []

    best_score = 0.0
    best_level = "none"
    signals: List[str] = []

    for _, rule in rules.iterrows():
        source_field = str(rule.get("source_field", "case_text"))
        trigger = str(rule.get("trigger", ""))
        match_type = str(rule.get("match_type", "contains")).lower()

        text = source_value(case, source_field)
        matched = contains_match(text, trigger) if match_type == "contains" else contains_match(text, trigger)

        if matched:
            score = to_float(rule.get(score_col, 0))
            reason = str(rule.get("reason", ""))
            signals.append(f"{trigger} ({reason}, Score {score})")

            if score > best_score:
                best_score = score
                if level_col and level_col in rules.columns:
                    best_level = str(rule.get(level_col, "none"))
                else:
                    best_level = "matched"

    return min(best_score, max_score), best_level, signals


def derive_fleet_value(case: pd.Series, rules: pd.DataFrame) -> Tuple[str, float, float, List[str]]:
    """
    Fleet-/Plural-Logik:
    - hÃ¶chsten fleet_value_score merken
    - hÃ¶chsten business_value_modifier merken
    - nicht addieren
    """
    if rules.empty:
        return "none", 0.0, 0.0, []

    best_fleet_score = 0.0
    best_modifier = 0.0
    best_level = "none"
    signals: List[str] = []

    for _, rule in rules.iterrows():
        source_field = str(rule.get("source_field", "case_text"))
        trigger = str(rule.get("trigger", ""))
        match_type = str(rule.get("match_type", "contains")).lower()
        text = source_value(case, source_field)
        matched = contains_match(text, trigger) if match_type == "contains" else contains_match(text, trigger)

        if matched:
            fleet_score = to_float(rule.get("fleet_value_score", 0))
            modifier = to_float(rule.get("business_value_modifier", 0))
            level = str(rule.get("fleet_signal_level", "none"))
            reason = str(rule.get("reason", ""))

            signals.append(f"{trigger} ({reason}, Fleet {fleet_score}, Modifier {modifier})")

            if modifier > best_modifier:
                best_modifier = modifier
            if fleet_score > best_fleet_score:
                best_fleet_score = fleet_score
                best_level = level

    return best_level, min(best_fleet_score, 10.0), min(best_modifier, 10.0), signals


def derive_customer_type(
    case: pd.Series,
    rules: pd.DataFrame,
    fleet_signal_level: str,
    fleet_signals: List[str]
) -> Tuple[str, str, List[str]]:
    """
    B2B/B2C:
    - stop_on_match=yes gewinnt sofort
    - sonst Scores sammeln
    - Fleet high/very_high wirkt als starker B2B-Indikator
    """
    b2b_score = 0.0
    b2c_score = 0.0
    signals: List[str] = []

    if not rules.empty:
        for _, rule in rules.iterrows():
            source_field = str(rule.get("source_field", "case_text"))
            trigger = str(rule.get("trigger", ""))
            match_type = str(rule.get("match_type", "contains")).lower()
            derived = str(rule.get("derived_customer_type", "")).strip()
            score = to_float(rule.get("score", 0))
            stop_on_match = str(rule.get("stop_on_match", "no")).lower()
            reason = str(rule.get("reason", ""))

            text = source_value(case, source_field)
            matched = contains_match(text, trigger) if match_type == "contains" else contains_match(text, trigger)

            if matched:
                signals.append(f"{trigger} â†’ {derived} ({reason}, Score {score})")

                if stop_on_match == "yes":
                    return derived, f"Starker Treffer: {trigger} / {reason}", signals

                if derived.upper() == "B2B":
                    b2b_score += score
                elif derived.upper() == "B2C":
                    b2c_score += score

    if fleet_signal_level in ["high", "very_high"]:
        b2b_score += 8
        signals.append(f"Fleet-/Plural-Signal als B2B-Indikator ({fleet_signal_level})")

    if b2b_score >= b2c_score + 3 and b2b_score > 0:
        return "B2B", f"B2B Score {b2b_score} > B2C Score {b2c_score}", signals
    if b2c_score >= b2b_score + 3 and b2c_score > 0:
        return "B2C", f"B2C Score {b2c_score} > B2B Score {b2b_score}", signals

    return "unknown", f"Keine eindeutige Ableitung B2B/B2C ({b2b_score}/{b2c_score})", signals


def derive_case_type(case: pd.Series, rules: pd.DataFrame) -> Tuple[str, str, List[str]]:
    if rules.empty:
        return "unknown", "Keine case_type_rules vorhanden", []

    best_score = -1.0
    best_type = "unknown"
    best_reason = "Kein Treffer"
    signals: List[str] = []

    for _, rule in rules.iterrows():
        source_field = str(rule.get("source_field", "case_text"))
        trigger = str(rule.get("trigger", ""))
        match_type = str(rule.get("match_type", "contains")).lower()
        derived = str(rule.get("derived_case_type", "unknown"))
        score = to_float(rule.get("score", 0))
        reason = str(rule.get("reason", ""))

        text = source_value(case, source_field)
        matched = contains_match(text, trigger) if match_type == "contains" else contains_match(text, trigger)

        if matched:
            signals.append(f"{trigger} â†’ {derived} ({reason}, Score {score})")
            if score > best_score:
                best_score = score
                best_type = derived
                best_reason = f"{trigger}: {reason}"

    return best_type, best_reason, signals


def history_score(previous_cases: Any) -> float:
    n = to_int(previous_cases, 0)
    if n <= 0:
        return 0.0
    if n == 1:
        return 2.0
    if n == 2:
        return 4.0
    if n == 3:
        return 6.0
    if n == 4:
        return 8.0
    return 10.0


def read_weights(df: pd.DataFrame) -> Dict[str, float]:
    defaults = {
        "business_value_score": 15.0,
        "urgency_score": 25.0,
        "escalation_score": 25.0,
        "emotion_score": 15.0,
        "history_score": 10.0,
        "claim_amount_score": 10.0,
    }

    if df.empty:
        return defaults

    component_col = None
    for possible in ["score_component", "component", "score_type"]:
        if possible in df.columns:
            component_col = possible
            break

    weight_col = None
    for possible in ["weight_percent", "weight", "weight_pct"]:
        if possible in df.columns:
            weight_col = possible
            break

    if not component_col or not weight_col:
        return defaults

    weights = defaults.copy()
    for _, row in df.iterrows():
        component = str(row.get(component_col, "")).strip()
        if component:
            weights[component] = to_float(row.get(weight_col, 0), defaults.get(component, 0))

    return weights


def weighted_priority_score(
    business_value_score: float,
    urgency_score: float,
    escalation_score: float,
    emotion_score: float,
    history_score_value: float,
    claim_amount_score: float,
    weights: Dict[str, float]
) -> float:
    # Jeder Teilscore ist 0-10. Gewicht ist Prozent. Ergebnis ist 0-100.
    score = (
        business_value_score * weights.get("business_value_score", 15.0)
        + urgency_score * weights.get("urgency_score", 25.0)
        + escalation_score * weights.get("escalation_score", 25.0)
        + emotion_score * weights.get("emotion_score", 15.0)
        + history_score_value * weights.get("history_score", 10.0)
        + claim_amount_score * weights.get("claim_amount_score", 10.0)
    ) / 10.0

    return round(min(max(score, 0.0), 100.0), 1)


def risk_level_from_scores(escalation_score: float, emotion_score: float) -> str:
    """
    MVP-Risiko:
    PrimÃ¤r Eskalation. Emotion kann leicht erhÃ¶hen, aber nicht allein rot erzeugen.
    """
    combined = max(escalation_score, min(10.0, escalation_score + emotion_score * 0.3))

    if combined >= 8:
        return "red"
    if combined >= 6:
        return "orange"
    if combined >= 3:
        return "yellow"
    return "green"


def priority_class_from_score(score: float) -> str:
    if score >= 76:
        return "P1"
    if score >= 56:
        return "P2"
    if score >= 31:
        return "P3"
    return "P4"


def apply_priority_floor(
    raw_priority_class: str,
    risk_level: str,
    derived_case_type: str,
    urgency_score: float,
    urgency_level: str,
    customer_type: str = "unknown",
    business_value_score: float = 0.0,
    history_score_value: float = 0.0,
) -> Tuple[str, str]:
    """
    MVP-Hard-Floors v3:
    - red risk = mindestens P1
    - Legal = mindestens P1
    - sehr hohe Dringlichkeit = mindestens P2
    - hoher Business Value + hohe Dringlichkeit = mindestens P2
    - B2B + orange risk + hoher Business Value = mindestens P2
    """
    final_priority = raw_priority_class
    reasons = []

    if risk_level == "red":
        final_priority = max_priority(final_priority, "P1")
        reasons.append("red risk â†’ MindestprioritÃ¤t P1")

    if normalize_text(derived_case_type) == "legal":
        final_priority = max_priority(final_priority, "P1")
        reasons.append("Legal Case Type â†’ MindestprioritÃ¤t P1")

    if urgency_score >= 9 or urgency_level == "very_high":
        final_priority = max_priority(final_priority, "P2")
        reasons.append("very high urgency â†’ MindestprioritÃ¤t P2")

    if business_value_score >= 8 and urgency_score >= 7:
        final_priority = max_priority(final_priority, "P2")
        reasons.append("high business value + high urgency â†’ MindestprioritÃ¤t P2")

    if normalize_text(customer_type) == "b2b" and risk_level == "orange" and business_value_score >= 8:
        final_priority = max_priority(final_priority, "P2")
        reasons.append("B2B + orange risk + high business value â†’ MindestprioritÃ¤t P2")

    if normalize_text(customer_type) == "b2b" and risk_level == "orange" and history_score_value >= 8:
        final_priority = max_priority(final_priority, "P2")
        reasons.append("B2B + orange risk + repeated history â†’ MindestprioritÃ¤t P2")

    return final_priority, " | ".join(reasons) if reasons else "none"


def business_level_from_score(score: float) -> str:
    if score >= 9:
        return "very_high"
    if score >= 7:
        return "high"
    if score >= 4:
        return "medium"
    if score > 0:
        return "low"
    return "unknown"


def urgency_level_from_score(score: float) -> str:
    if score >= 9:
        return "very_high"
    if score >= 7:
        return "high"
    if score >= 4:
        return "medium"
    if score > 0:
        return "low"
    return "none"


def apply_template_placeholders(text: str, case: pd.Series) -> str:
    """Ersetzt einfache Platzhalter im empfohlenen Antworttext."""
    if text is None:
        return ""

    result = str(text)
    replacements = {
        "[Kundenname]": str(case.get("customer_name", "")).strip() or "Kundenname",
        "[Fahrzeugmodell]": str(case.get("vehicle_model", "")).strip() or "Fahrzeugmodell",
        "[FIN]": str(case.get("vin", "")).strip() or "FIN",
        "[Thema/Fallbezug]": str(case.get("case_subject", "")).strip() or "Ihr Anliegen",
    }

    for placeholder, value in replacements.items():
        result = result.replace(placeholder, value)

    return result


def _match_value(rule_value: str, actual_value: str) -> bool:
    rv = normalize_text(rule_value)
    av = normalize_text(actual_value)
    return rv in ["", "any", "*", "alle"] or rv == av


def _specificity(rule_value: str) -> int:
    rv = normalize_text(rule_value)
    return 0 if rv in ["", "any", "*", "alle"] else 1


def choose_template_from_master(
    template_master: pd.DataFrame,
    case_type: str,
    customer_type: str,
    risk_level: str,
    priority_class: str,
    case: pd.Series,
) -> Dict[str, str]:
    """
    Neue MVP-v7-Logik:
    - 8 Master-Templates im Sheet template_master
    - Red Risk oder Legal Ã¼berschreibt alles
    - Danach Matching auf case_type/customer_type/risk/priority
    - Wildcards: any, *, leer
    - Sortierung: niedriger selection_rank zuerst, dann SpezifitÃ¤t
    """
    empty_result = {
        "recommended_template_id": "NO_TEMPLATE",
        "recommended_next_action": "Manuell prÃ¼fen",
        "agent_warning": "Kein template_master vorhanden",
        "tone_level": "",
        "deescalation_phrases": "",
        "forbidden_claims": "",
        "recommended_customer_reply": "",
        "template_selection_reason": "template_master fehlt oder ist leer",
    }

    if template_master.empty:
        return empty_result

    df = template_master.copy().fillna("")

    # erwartete Spalten nachziehen, falls einzelne fehlen
    required_cols = [
        "template_id",
        "case_type_match",
        "customer_type_match",
        "risk_match",
        "priority_match",
        "selection_rank",
        "tone_level",
        "recommended_next_action",
        "agent_warning",
        "deescalation_phrases",
        "forbidden_claims",
        "recommended_customer_reply",
    ]
    for col in required_cols:
        if col not in df.columns:
            df[col] = ""

    norm_case_type = normalize_text(case_type)
    norm_risk = normalize_text(risk_level)

    # 1. Hard override: red risk oder Legal Case Type
    if norm_risk == "red" or norm_case_type == "legal":
        legal = df[df["template_id"].astype(str).apply(normalize_text) == "legal_red_p1_01"]
        if not legal.empty:
            row = legal.iloc[0]
            reply = apply_template_placeholders(row.get("recommended_customer_reply", ""), case)
            return {
                "recommended_template_id": str(row.get("template_id", "LEGAL_RED_P1_01")),
                "recommended_next_action": str(row.get("recommended_next_action", "Senior/Legal Review einbinden")),
                "agent_warning": str(row.get("agent_warning", "Keine Zusagen ohne PrÃ¼fung")),
                "tone_level": str(row.get("tone_level", "sachlich-rechtssicher")),
                "deescalation_phrases": str(row.get("deescalation_phrases", "")),
                "forbidden_claims": str(row.get("forbidden_claims", "")),
                "recommended_customer_reply": reply,
                "template_selection_reason": "Hard override: red risk oder Legal Case Type",
            }

    def candidate_rows(relax_priority=False, relax_risk=False, relax_customer=False, relax_case=False):
        rows = []
        for _, row in df.iterrows():
            case_ok = True if relax_case else _match_value(row.get("case_type_match", ""), case_type)
            cust_ok = True if relax_customer else _match_value(row.get("customer_type_match", ""), customer_type)
            risk_ok = True if relax_risk else _match_value(row.get("risk_match", ""), risk_level)
            prio_ok = True if relax_priority else _match_value(row.get("priority_match", ""), priority_class)

            if case_ok and cust_ok and risk_ok and prio_ok:
                rows.append(row)

        return pd.DataFrame(rows)

    # 2. Matching-Kaskade
    cascades = [
        ("exact_or_any_match", {}),
        # Erst Risk relaxen, damit P2-PrioritÃ¤t stÃ¤rker bleibt als ein P3-Template mit any-risk.
        ("relax_risk", {"relax_risk": True}),
        ("relax_priority", {"relax_priority": True}),
        ("relax_priority_and_risk", {"relax_priority": True, "relax_risk": True}),
        ("relax_priority_risk_customer", {"relax_priority": True, "relax_risk": True, "relax_customer": True}),
        ("fallback_any", {"relax_priority": True, "relax_risk": True, "relax_customer": True, "relax_case": True}),
    ]

    for reason, kwargs in cascades:
        cand = candidate_rows(**kwargs)
        if not cand.empty:
            # SpezifitÃ¤t bewerten: exakte Regeln vor any.
            # Wichtig: Das generische INFO-Template darf erst im echten Fallback greifen,
            # sonst verhindert es bessere Relaxed Matches wie Reklamation/B2C/P2.
            # LEGAL_RED_P1_01 darf auÃŸerhalb des Hard Overrides nicht matchen.
            cand = cand.copy()
            cand = cand[cand["template_id"].astype(str).apply(normalize_text) != "legal_red_p1_01"].copy()
            if cand.empty:
                continue
            cand["_rank"] = cand["selection_rank"].apply(to_float)
            cand["_rank"] = cand["_rank"].replace(0, 999)
            cand["_specificity"] = (
                cand["case_type_match"].apply(_specificity)
                + cand["customer_type_match"].apply(_specificity)
                + cand["risk_match"].apply(_specificity)
                + cand["priority_match"].apply(_specificity)
            )

            if reason != "fallback_any":
                cand_specific = cand[cand["_specificity"] > 0].copy()
                if cand_specific.empty:
                    continue
                cand = cand_specific

            cand = cand.sort_values(by=["_rank", "_specificity"], ascending=[True, False])
            row = cand.iloc[0]

            reply = apply_template_placeholders(row.get("recommended_customer_reply", ""), case)

            return {
                "recommended_template_id": str(row.get("template_id", "NO_TEMPLATE")),
                "recommended_next_action": str(row.get("recommended_next_action", "Manuell prÃ¼fen")),
                "agent_warning": str(row.get("agent_warning", "")),
                "tone_level": str(row.get("tone_level", "")),
                "deescalation_phrases": str(row.get("deescalation_phrases", "")),
                "forbidden_claims": str(row.get("forbidden_claims", "")),
                "recommended_customer_reply": reply,
                "template_selection_reason": reason,
            }

    return empty_result


def choose_template(
    template_mapping: pd.DataFrame,
    case_type: str,
    customer_type: str,
    risk_level: str,
    priority_class: str
) -> Tuple[str, str, str]:
    """
    Legacy-Fallback fÃ¼r alte Rulebooks ohne template_master.
    """
    if template_mapping.empty:
        return "NO_TEMPLATE", "Manuell prÃ¼fen", "Kein template_mapping vorhanden"

    df = template_mapping.copy()
    for col in ["derived_case_type", "case_type"]:
        if col in df.columns:
            df["_case_type"] = df[col]
            break
    if "_case_type" not in df.columns:
        df["_case_type"] = ""

    for col in ["derived_customer_type", "customer_type"]:
        if col in df.columns:
            df["_customer_type"] = df[col]
            break
    if "_customer_type" not in df.columns:
        df["_customer_type"] = ""

    def norm_col(col: str) -> pd.Series:
        return df[col].apply(normalize_text)

    ct = normalize_text(case_type)
    cust = normalize_text(customer_type)
    risk = normalize_text(risk_level)
    prio = normalize_text(priority_class)

    candidates = [
        df[
            (norm_col("_case_type") == ct)
            & (norm_col("_customer_type") == cust)
            & (df.get("risk_level", "").astype(str).apply(normalize_text) == risk)
            & (df.get("priority_class", "").astype(str).apply(normalize_text) == prio)
        ],
        df[
            (norm_col("_case_type") == ct)
            & (df.get("risk_level", "").astype(str).apply(normalize_text) == risk)
            & (df.get("priority_class", "").astype(str).apply(normalize_text) == prio)
        ],
        df[(norm_col("_case_type") == ct) & (norm_col("_customer_type") == cust)],
        df[norm_col("_case_type") == ct],
        df,
    ]

    for cand in candidates:
        if not cand.empty:
            row = cand.iloc[0]
            template_id = str(
                row.get("template_id", row.get("recommended_template_id", "NO_TEMPLATE"))
            )
            action = str(row.get("recommended_next_action", row.get("recommended_strategy", "Manuell prÃ¼fen")))
            warning = str(row.get("agent_warning", row.get("warning", "")))
            return template_id, action, warning

    return "NO_TEMPLATE", "Manuell prÃ¼fen", "Kein passendes Template gefunden"




# ------------------------------------------------------------
# Excel Output Styling
# ------------------------------------------------------------

def _safe_sheet(wb, name: str):
    return wb[name] if name in wb.sheetnames else None


def _col_index_by_header(ws) -> Dict[str, int]:
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value)] = cell.column
    return headers


def _apply_basic_sheet_style(ws, header_fill="#0F2742"):
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=header_fill.replace("#", ""))
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            cell.font = Font(size=9, color="111827")

    ws.row_dimensions[1].height = 34


def _set_widths(ws, widths: Dict[str, int], default_width: int = 16, max_width: int = 48):
    headers = _col_index_by_header(ws)
    for col_name, idx in headers.items():
        width = widths.get(col_name, default_width)
        ws.column_dimensions[get_column_letter(idx)].width = min(width, max_width)


def _apply_priority_and_risk_formatting(ws):
    """
    Direkte Zellformatierung statt Conditional Formatting.
    Grund: Excel zeigte bei P1/P2/red teils weiÃŸe Schrift ohne sichtbare FÃ¼llung.
    """
    headers = _col_index_by_header(ws)

    priority_col = headers.get("case_priority_class")
    risk_col = headers.get("escalation_risk_level")
    urgency_col = headers.get("urgency_level")
    bv_col = headers.get("business_value_level")
    score_col = headers.get("priority_score")

    priority_styles = {
        "P1": ("7F1D1D", "FFFFFF"),
        "P2": ("F97316", "FFFFFF"),
        "P3": ("FACC15", "111827"),
        "P4": ("D1FAE5", "065F46"),
    }

    risk_styles = {
        "red": ("991B1B", "FFFFFF"),
        "orange": ("FDBA74", "7C2D12"),
        "yellow": ("FEF08A", "713F12"),
        "green": ("BBF7D0", "14532D"),
    }

    urgency_styles = {
        "very_high": ("FCA5A5", "7F1D1D"),
        "high": ("FED7AA", "7C2D12"),
        "medium": ("FEF3C7", "713F12"),
        "low": ("E0F2FE", "075985"),
        "none": ("F8FAFC", "475569"),
    }

    bv_styles = {
        "very_high": ("C4B5FD", "2E1065"),
        "high": ("DDD6FE", "4C1D95"),
        "medium": ("E0E7FF", "3730A3"),
        "low": ("E2E8F0", "334155"),
        "unknown": ("F8FAFC", "475569"),
    }

    def style_cell(cell, styles):
        value = str(cell.value).strip() if cell.value is not None else ""
        if value in styles:
            fill, font = styles[value]
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(color=font, bold=True, size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(2, ws.max_row + 1):
        if priority_col:
            style_cell(ws.cell(row, priority_col), priority_styles)
        if risk_col:
            style_cell(ws.cell(row, risk_col), risk_styles)
        if urgency_col:
            style_cell(ws.cell(row, urgency_col), urgency_styles)
        if bv_col:
            style_cell(ws.cell(row, bv_col), bv_styles)
        if score_col:
            cell = ws.cell(row, score_col)
            try:
                score = float(cell.value)
                cell.number_format = "0.0"
                if score >= 76:
                    cell.fill = PatternFill("solid", fgColor="FCA5A5")
                    cell.font = Font(color="7F1D1D", bold=True, size=9)
                elif score >= 56:
                    cell.fill = PatternFill("solid", fgColor="FED7AA")
                    cell.font = Font(color="7C2D12", bold=True, size=9)
                elif score >= 31:
                    cell.fill = PatternFill("solid", fgColor="FEF3C7")
                    cell.font = Font(color="713F12", bold=True, size=9)
                else:
                    cell.fill = PatternFill("solid", fgColor="DBEAFE")
                    cell.font = Font(color="1E3A8A", bold=True, size=9)
            except Exception:
                pass


def _style_agent_view(ws):
    _apply_basic_sheet_style(ws, "#7F1D1D")
    widths = {
        "case_id": 14,
        "case_priority_class": 14,
        "escalation_risk_level": 16,
        "priority_score": 13,
        "urgency_level": 16,
        "business_value_level": 18,
        "customer_name": 22,
        "vehicle_model": 16,
        "recommended_template_id": 32,
        "tone_level": 22,
        "recommended_next_action": 44,
        "agent_warning": 44,
        "deescalation_phrases": 46,
        "forbidden_claims": 46,
        "recommended_customer_reply": 70,
        "template_selection_reason": 24,
        "decision_reason": 58,
    }
    _set_widths(ws, widths, default_width=18, max_width=58)
    _apply_priority_and_risk_formatting(ws)

    # Agent-View visuell stÃ¤rker gewichten
    for col in range(1, ws.max_column + 1):
        ws.cell(1, col).font = Font(color="FFFFFF", bold=True, size=10)

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 48


def _style_output_simulation(ws):
    _apply_basic_sheet_style(ws, "#0F2742")
    headers = _col_index_by_header(ws)

    input_cols = {
        "case_id", "customer_name", "vehicle_model", "vin", "case_subject",
        "case_text", "previous_cases", "language"
    }
    reasoning_cols = {
        "derived_customer_type", "customer_type_reason", "derived_case_type",
        "case_type_reason", "base_business_value_score", "base_business_value_level",
        "fleet_signal_level", "fleet_value_score", "fleet_value_modifier",
        "final_business_value_score", "business_value_level", "urgency_score",
        "urgency_level", "escalation_score", "emotion_score", "history_score",
        "claim_amount_score", "priority_score", "raw_priority_class",
        "priority_floor_applied", "detected_signals"
    }
    agent_cols = {
        "case_priority_class", "escalation_risk_level", "recommended_template_id",
        "tone_level", "recommended_next_action", "agent_warning", "deescalation_phrases",
        "forbidden_claims", "recommended_customer_reply", "template_selection_reason", "decision_reason"
    }

    for name, idx in headers.items():
        cell = ws.cell(1, idx)
        if name in input_cols:
            cell.fill = PatternFill("solid", fgColor="1D4E89")
        elif name in reasoning_cols:
            cell.fill = PatternFill("solid", fgColor="6D28D9")
        elif name in agent_cols:
            cell.fill = PatternFill("solid", fgColor="991B1B")

    widths = {
        "case_id": 12,
        "customer_name": 22,
        "vehicle_model": 16,
        "vin": 18,
        "case_subject": 28,
        "case_text": 52,
        "previous_cases": 13,
        "language": 10,
        "derived_customer_type": 18,
        "customer_type_reason": 38,
        "derived_case_type": 22,
        "case_type_reason": 34,
        "base_business_value_score": 16,
        "fleet_signal_level": 16,
        "fleet_value_score": 15,
        "fleet_value_modifier": 16,
        "final_business_value_score": 18,
        "business_value_level": 18,
        "urgency_score": 14,
        "urgency_level": 16,
        "escalation_score": 16,
        "emotion_score": 14,
        "history_score": 14,
        "priority_score": 14,
        "priority_floor_applied": 42,
        "detected_signals": 64,
        "case_priority_class": 15,
        "escalation_risk_level": 18,
        "recommended_template_id": 34,
        "tone_level": 22,
        "recommended_next_action": 48,
        "agent_warning": 48,
        "deescalation_phrases": 52,
        "forbidden_claims": 52,
        "recommended_customer_reply": 72,
        "template_selection_reason": 28,
        "decision_reason": 64,
    }
    _set_widths(ws, widths, default_width=16, max_width=64)
    _apply_priority_and_risk_formatting(ws)

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 54


def _create_dashboard(wb):
    if "Dashboard_MVP" in wb.sheetnames:
        del wb["Dashboard_MVP"]

    ws = wb.create_sheet("Dashboard_MVP", 0)
    ws.sheet_view.showGridLines = False

    agent_ws = wb["Agent_View_MVP"] if "Agent_View_MVP" in wb.sheetnames else None

    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "AI Case Engine MVP Dashboard"
    ws["A1"].fill = PatternFill("solid", fgColor="0F2742")
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws["A2"] = "Automatisch erzeugte Management- und AgentenÃ¼bersicht"
    ws["A2"].font = Font(color="475569", italic=True)
    ws.merge_cells("A2:H2")

    # Agent View in Records lesen, damit Dashboard statische Werte statt Excel-Formeln nutzt
    records = []
    if agent_ws:
        headers = [cell.value for cell in agent_ws[1]]
        for row in agent_ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None and str(v).strip() != "" for v in row):
                records.append(dict(zip(headers, row)))

    def count_where(field, value):
        return sum(1 for r in records if str(r.get(field, "")).lower() == str(value).lower())

    def count_bv_high():
        return sum(
            1 for r in records
            if str(r.get("business_value_level", "")).lower() in ["high", "very_high"]
        )

    # KPI cards
    cards = [
        ("A4", "Cases gesamt", len(records), "#DBEAFE"),
        ("C4", "P1 Cases", count_where("case_priority_class", "P1"), "#FECACA"),
        ("E4", "P2 Cases", count_where("case_priority_class", "P2"), "#FED7AA"),
        ("G4", "Red Risk", count_where("escalation_risk_level", "red"), "#FCA5A5"),
        ("A7", "Orange Risk", count_where("escalation_risk_level", "orange"), "#FDBA74"),
        ("C7", "Very High Urgency", count_where("urgency_level", "very_high"), "#FEF3C7"),
        ("E7", "High Business Value", count_bv_high(), "#EDE9FE"),
        ("G7", "Run Status", "OK", "#DCFCE7"),
    ]

    for cell_ref, label, value, fill in cards:
        col = ws[cell_ref].column
        row = ws[cell_ref].row
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        ws.cell(row, col).value = label
        ws.cell(row, col).fill = PatternFill("solid", fgColor=fill.replace("#", ""))
        ws.cell(row, col).font = Font(color="111827", bold=True, size=10)
        ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row + 1, col).value = value
        ws.cell(row + 1, col).fill = PatternFill("solid", fgColor=fill.replace("#", ""))
        ws.cell(row + 1, col).font = Font(color="111827", bold=True, size=20)
        ws.cell(row + 1, col).alignment = Alignment(horizontal="center", vertical="center")

    # Top Cases
    ws["A11"] = "Top priorisierte Cases"
    ws["A11"].font = Font(bold=True, size=13, color="0F172A")
    headers = ["Case", "Prio", "Risk", "Score", "Urgency", "Business Value", "Kunde", "Empfohlene Aktion"]
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(12, idx)
        cell.value = h
        cell.fill = PatternFill("solid", fgColor="0F2742")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Statische Werte aus Agent_View schreiben, keine Formeln mehr
    fields = [
        "case_id",
        "case_priority_class",
        "escalation_risk_level",
        "priority_score",
        "urgency_level",
        "business_value_level",
        "customer_name",
        "recommended_next_action",
    ]

    for row_offset, rec in enumerate(records[:10], start=13):
        for col_idx, field in enumerate(fields, start=1):
            cell = ws.cell(row_offset, col_idx)
            cell.value = rec.get(field, "")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=Side(style="thin", color="D9E2EC"),
                right=Side(style="thin", color="D9E2EC"),
                top=Side(style="thin", color="D9E2EC"),
                bottom=Side(style="thin", color="D9E2EC")
            )

    ws.freeze_panes = "A12"

    widths = {
        "A": 14, "B": 10, "C": 12, "D": 12, "E": 16, "F": 18, "G": 22, "H": 52
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for r in range(13, 23):
        ws.row_dimensions[r].height = 42

    # Direkte Formatierung auf Top Cases
    prio_styles = {
        "P1": ("7F1D1D", "FFFFFF"),
        "P2": ("F97316", "FFFFFF"),
        "P3": ("FACC15", "111827"),
        "P4": ("D1FAE5", "065F46"),
    }
    risk_styles = {
        "red": ("991B1B", "FFFFFF"),
        "orange": ("FDBA74", "7C2D12"),
        "yellow": ("FEF08A", "713F12"),
        "green": ("BBF7D0", "14532D"),
    }

    for r in range(13, 23):
        prio = str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value is not None else ""
        if prio in prio_styles:
            fill, font = prio_styles[prio]
            ws.cell(r, 2).fill = PatternFill("solid", fgColor=fill)
            ws.cell(r, 2).font = Font(color=font, bold=True)
            ws.cell(r, 2).alignment = Alignment(horizontal="center", vertical="center")

        risk = str(ws.cell(r, 3).value).strip() if ws.cell(r, 3).value is not None else ""
        if risk in risk_styles:
            fill, font = risk_styles[risk]
            ws.cell(r, 3).fill = PatternFill("solid", fgColor=fill)
            ws.cell(r, 3).font = Font(color=font, bold=True)
            ws.cell(r, 3).alignment = Alignment(horizontal="center", vertical="center")

        score_cell = ws.cell(r, 4)
        try:
            score = float(score_cell.value)
            score_cell.number_format = "0.0"
            if score >= 76:
                score_cell.fill = PatternFill("solid", fgColor="FCA5A5")
            elif score >= 56:
                score_cell.fill = PatternFill("solid", fgColor="FED7AA")
            elif score >= 31:
                score_cell.fill = PatternFill("solid", fgColor="FEF3C7")
            else:
                score_cell.fill = PatternFill("solid", fgColor="DBEAFE")
            score_cell.font = Font(color="111827", bold=True)
        except Exception:
            pass


def format_output_workbook(output_file: Path) -> None:
    wb = load_workbook(output_file)

    # Reorder key sheets after dashboard creation
    sim = _safe_sheet(wb, "Output_Simulation_MVP")
    agent = _safe_sheet(wb, "Agent_View_MVP")
    runinfo = _safe_sheet(wb, "Run_Info")

    if sim:
        _style_output_simulation(sim)
    if agent:
        _style_agent_view(agent)
    if runinfo:
        _apply_basic_sheet_style(runinfo, "#334155")
        _set_widths(runinfo, {"key": 24, "value": 90}, default_width=30, max_width=90)

    _create_dashboard(wb)

    # Tab colors
    if "Dashboard_MVP" in wb.sheetnames:
        wb["Dashboard_MVP"].sheet_properties.tabColor = "0F2742"
    if agent:
        agent.sheet_properties.tabColor = "991B1B"
    if sim:
        sim.sheet_properties.tabColor = "6D28D9"
    if runinfo:
        runinfo.sheet_properties.tabColor = "334155"

    wb.save(output_file)


def analyze_cases(cases: pd.DataFrame, rules: Dict[str, pd.DataFrame]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Run the exact MVP v7 case analysis for already loaded Streamlit data."""
    missing = [c for c in REQUIRED_INPUT_COLUMNS if c not in cases.columns]
    if missing:
        raise ValueError("Folgende Input-Spalten fehlen: " + ", ".join(missing))

    weights = read_weights(rules["score_weights"])

    output_rows = []

    for _, case in cases.iterrows():
        # 1. Business Value aus Modell
        base_bv_score, base_bv_level, base_bv_reason = derive_business_value(
            case, rules["business_value_matrix"]
        )

        # 2. Fleet Value / Plural / Flottenkontext
        fleet_level, fleet_score, fleet_modifier, fleet_signals = derive_fleet_value(
            case, rules["fleet_value_rules"]
        )

        final_bv_score = min(base_bv_score + fleet_modifier, 10.0)
        final_bv_level = business_level_from_score(final_bv_score)

        # 3. Kundentyp ableiten
        customer_type, customer_type_reason, customer_type_signals = derive_customer_type(
            case, rules["customer_type_rules"], fleet_level, fleet_signals
        )

        # 4. Case Type ableiten
        case_type, case_type_reason, case_type_signals = derive_case_type(
            case, rules["case_type_rules"]
        )

        # 5. Scores
        urgency_score, urgency_level_rule, urgency_signals = apply_keyword_rules(
            case,
            rules["urgency_keywords"],
            score_col="urgency_score",
            level_col="urgency_level",
            max_score=10.0,
        )
        urgency_level = urgency_level_rule if urgency_level_rule != "none" else urgency_level_from_score(urgency_score)

        escalation_score, escalation_level_rule, escalation_signals = apply_keyword_rules(
            case,
            rules["escalation_keywords"],
            score_col="escalation_score",
            level_col="risk_level",
            max_score=10.0,
        )

        emotion_score, emotion_level_rule, emotion_signals = apply_keyword_rules(
            case,
            rules["emotion_keywords"],
            score_col="emotion_score",
            level_col="emotion_level",
            max_score=10.0,
        )

        hist_score = history_score(case.get("previous_cases", 0))
        claim_amount_score = 0.0  # MVP Step 1: noch keine Betrags-Extraktion

        # 6. Priority Score / Risk / Priority Class
        priority_score = weighted_priority_score(
            business_value_score=final_bv_score,
            urgency_score=urgency_score,
            escalation_score=escalation_score,
            emotion_score=emotion_score,
            history_score_value=hist_score,
            claim_amount_score=claim_amount_score,
            weights=weights,
        )

        risk_level = risk_level_from_scores(escalation_score, emotion_score)
        raw_priority_class = priority_class_from_score(priority_score)
        final_priority_class, priority_floor_applied = apply_priority_floor(
            raw_priority_class,
            risk_level,
            case_type,
            urgency_score,
            urgency_level,
            customer_type=customer_type,
            business_value_score=final_bv_score,
            history_score_value=hist_score,
        )

        # 7. Template / Agent Action
        if not rules.get("template_master", pd.DataFrame()).empty:
            template_result = choose_template_from_master(
                rules["template_master"],
                case_type,
                customer_type,
                risk_level,
                final_priority_class,
                case,
            )
            template_id = template_result["recommended_template_id"]
            recommended_next_action = template_result["recommended_next_action"]
            agent_warning = template_result["agent_warning"]
            tone_level = template_result["tone_level"]
            deescalation_phrases = template_result["deescalation_phrases"]
            forbidden_claims = template_result["forbidden_claims"]
            recommended_customer_reply = template_result["recommended_customer_reply"]
            template_selection_reason = template_result["template_selection_reason"]
        else:
            template_id, recommended_next_action, agent_warning = choose_template(
                rules["template_mapping"],
                case_type,
                customer_type,
                risk_level,
                final_priority_class,
            )
            tone_level = ""
            deescalation_phrases = ""
            forbidden_claims = ""
            recommended_customer_reply = ""
            template_selection_reason = "legacy template_mapping"

        # 8. Reasoning
        detected_signals = []
        detected_signals.extend([f"Fleet: {s}" for s in fleet_signals])
        detected_signals.extend([f"CustomerType: {s}" for s in customer_type_signals])
        detected_signals.extend([f"CaseType: {s}" for s in case_type_signals])
        detected_signals.extend([f"Urgency: {s}" for s in urgency_signals])
        detected_signals.extend([f"Escalation: {s}" for s in escalation_signals])
        detected_signals.extend([f"Emotion: {s}" for s in emotion_signals])

        decision_reason = (
            f"Modellwert {case.get('vehicle_model', '')}: {base_bv_score}/10; "
            f"Fleet Modifier: +{fleet_modifier}; final Business Value: {final_bv_score}/10; "
            f"Urgency: {urgency_score}/10; Escalation: {escalation_score}/10; "
            f"Emotion: {emotion_score}/10; History: {hist_score}/10; "
            f"Risk: {risk_level}; Priority: {final_priority_class}."
        )

        # 9. Output Row
        output_rows.append({
            # Input
            "case_id": case.get("case_id", ""),
            "customer_name": case.get("customer_name", ""),
            "vehicle_model": case.get("vehicle_model", ""),
            "vin": case.get("vin", ""),
            "case_subject": case.get("case_subject", ""),
            "case_text": case.get("case_text", ""),
            "previous_cases": case.get("previous_cases", ""),
            "language": case.get("language", ""),

            # Reasoning
            "derived_customer_type": customer_type,
            "customer_type_reason": customer_type_reason,
            "derived_case_type": case_type,
            "case_type_reason": case_type_reason,
            "base_business_value_score": base_bv_score,
            "base_business_value_level": base_bv_level,
            "fleet_signal_level": fleet_level,
            "fleet_value_score": fleet_score,
            "fleet_value_modifier": fleet_modifier,
            "final_business_value_score": final_bv_score,
            "business_value_level": final_bv_level,
            "urgency_score": urgency_score,
            "urgency_level": urgency_level,
            "escalation_score": escalation_score,
            "emotion_score": emotion_score,
            "history_score": hist_score,
            "claim_amount_score": claim_amount_score,
            "priority_score": priority_score,
            "raw_priority_class": raw_priority_class,
            "priority_floor_applied": priority_floor_applied,
            "detected_signals": " | ".join(detected_signals),

            # Agent Output
            "case_priority_class": final_priority_class,
            "escalation_risk_level": risk_level,
            "recommended_template_id": template_id,
            "tone_level": tone_level,
            "recommended_next_action": recommended_next_action,
            "agent_warning": agent_warning,
            "deescalation_phrases": deescalation_phrases,
            "forbidden_claims": forbidden_claims,
            "recommended_customer_reply": recommended_customer_reply,
            "template_selection_reason": template_selection_reason,
            "decision_reason": decision_reason,
        })

    out = pd.DataFrame(output_rows)

    # technische Sortierspalten
    out["_priority_sort"] = out["case_priority_class"].apply(priority_rank)
    out["_risk_sort"] = out["escalation_risk_level"].apply(risk_rank)
    out = out.sort_values(
        by=["_priority_sort", "_risk_sort", "priority_score"],
        ascending=[True, True, False]
    ).drop(columns=["_priority_sort", "_risk_sort"])

    agent_cols = [
        "case_id",
        "case_priority_class",
        "escalation_risk_level",
        "priority_score",
        "urgency_level",
        "business_value_level",
        "customer_name",
        "vehicle_model",
        "recommended_template_id",
        "tone_level",
        "recommended_next_action",
        "agent_warning",
        "deescalation_phrases",
        "forbidden_claims",
        "recommended_customer_reply",
        "template_selection_reason",
        "decision_reason",
    ]

    agent_view = out[[c for c in agent_cols if c in out.columns]].copy()
    return out, agent_view

# ------------------------------------------------------------
# Main Engine
# ------------------------------------------------------------

def run_engine() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    input_file = find_input_file()
    rulebook_file = find_rulebook_file()

    log(f"Input-Datei: {input_file}")
    log(f"Rulebook:    {rulebook_file}")

    cases = read_cases(input_file)
    rules = read_rulebook(rulebook_file)
    weights = read_weights(rules["score_weights"])

    output_rows = []

    for _, case in cases.iterrows():
        # 1. Business Value aus Modell
        base_bv_score, base_bv_level, base_bv_reason = derive_business_value(
            case, rules["business_value_matrix"]
        )

        # 2. Fleet Value / Plural / Flottenkontext
        fleet_level, fleet_score, fleet_modifier, fleet_signals = derive_fleet_value(
            case, rules["fleet_value_rules"]
        )

        final_bv_score = min(base_bv_score + fleet_modifier, 10.0)
        final_bv_level = business_level_from_score(final_bv_score)

        # 3. Kundentyp ableiten
        customer_type, customer_type_reason, customer_type_signals = derive_customer_type(
            case, rules["customer_type_rules"], fleet_level, fleet_signals
        )

        # 4. Case Type ableiten
        case_type, case_type_reason, case_type_signals = derive_case_type(
            case, rules["case_type_rules"]
        )

        # 5. Scores
        urgency_score, urgency_level_rule, urgency_signals = apply_keyword_rules(
            case,
            rules["urgency_keywords"],
            score_col="urgency_score",
            level_col="urgency_level",
            max_score=10.0,
        )
        urgency_level = urgency_level_rule if urgency_level_rule != "none" else urgency_level_from_score(urgency_score)

        escalation_score, escalation_level_rule, escalation_signals = apply_keyword_rules(
            case,
            rules["escalation_keywords"],
            score_col="escalation_score",
            level_col="risk_level",
            max_score=10.0,
        )

        emotion_score, emotion_level_rule, emotion_signals = apply_keyword_rules(
            case,
            rules["emotion_keywords"],
            score_col="emotion_score",
            level_col="emotion_level",
            max_score=10.0,
        )

        hist_score = history_score(case.get("previous_cases", 0))
        claim_amount_score = 0.0  # MVP Step 1: noch keine Betrags-Extraktion

        # 6. Priority Score / Risk / Priority Class
        priority_score = weighted_priority_score(
            business_value_score=final_bv_score,
            urgency_score=urgency_score,
            escalation_score=escalation_score,
            emotion_score=emotion_score,
            history_score_value=hist_score,
            claim_amount_score=claim_amount_score,
            weights=weights,
        )

        risk_level = risk_level_from_scores(escalation_score, emotion_score)
        raw_priority_class = priority_class_from_score(priority_score)
        final_priority_class, priority_floor_applied = apply_priority_floor(
            raw_priority_class,
            risk_level,
            case_type,
            urgency_score,
            urgency_level,
            customer_type=customer_type,
            business_value_score=final_bv_score,
            history_score_value=hist_score,
        )

        # 7. Template / Agent Action
        if not rules.get("template_master", pd.DataFrame()).empty:
            template_result = choose_template_from_master(
                rules["template_master"],
                case_type,
                customer_type,
                risk_level,
                final_priority_class,
                case,
            )
            template_id = template_result["recommended_template_id"]
            recommended_next_action = template_result["recommended_next_action"]
            agent_warning = template_result["agent_warning"]
            tone_level = template_result["tone_level"]
            deescalation_phrases = template_result["deescalation_phrases"]
            forbidden_claims = template_result["forbidden_claims"]
            recommended_customer_reply = template_result["recommended_customer_reply"]
            template_selection_reason = template_result["template_selection_reason"]
        else:
            template_id, recommended_next_action, agent_warning = choose_template(
                rules["template_mapping"],
                case_type,
                customer_type,
                risk_level,
                final_priority_class,
            )
            tone_level = ""
            deescalation_phrases = ""
            forbidden_claims = ""
            recommended_customer_reply = ""
            template_selection_reason = "legacy template_mapping"

        # 8. Reasoning
        detected_signals = []
        detected_signals.extend([f"Fleet: {s}" for s in fleet_signals])
        detected_signals.extend([f"CustomerType: {s}" for s in customer_type_signals])
        detected_signals.extend([f"CaseType: {s}" for s in case_type_signals])
        detected_signals.extend([f"Urgency: {s}" for s in urgency_signals])
        detected_signals.extend([f"Escalation: {s}" for s in escalation_signals])
        detected_signals.extend([f"Emotion: {s}" for s in emotion_signals])

        decision_reason = (
            f"Modellwert {case.get('vehicle_model', '')}: {base_bv_score}/10; "
            f"Fleet Modifier: +{fleet_modifier}; final Business Value: {final_bv_score}/10; "
            f"Urgency: {urgency_score}/10; Escalation: {escalation_score}/10; "
            f"Emotion: {emotion_score}/10; History: {hist_score}/10; "
            f"Risk: {risk_level}; Priority: {final_priority_class}."
        )

        # 9. Output Row
        output_rows.append({
            # Input
            "case_id": case.get("case_id", ""),
            "customer_name": case.get("customer_name", ""),
            "vehicle_model": case.get("vehicle_model", ""),
            "vin": case.get("vin", ""),
            "case_subject": case.get("case_subject", ""),
            "case_text": case.get("case_text", ""),
            "previous_cases": case.get("previous_cases", ""),
            "language": case.get("language", ""),

            # Reasoning
            "derived_customer_type": customer_type,
            "customer_type_reason": customer_type_reason,
            "derived_case_type": case_type,
            "case_type_reason": case_type_reason,
            "base_business_value_score": base_bv_score,
            "base_business_value_level": base_bv_level,
            "fleet_signal_level": fleet_level,
            "fleet_value_score": fleet_score,
            "fleet_value_modifier": fleet_modifier,
            "final_business_value_score": final_bv_score,
            "business_value_level": final_bv_level,
            "urgency_score": urgency_score,
            "urgency_level": urgency_level,
            "escalation_score": escalation_score,
            "emotion_score": emotion_score,
            "history_score": hist_score,
            "claim_amount_score": claim_amount_score,
            "priority_score": priority_score,
            "raw_priority_class": raw_priority_class,
            "priority_floor_applied": priority_floor_applied,
            "detected_signals": " | ".join(detected_signals),

            # Agent Output
            "case_priority_class": final_priority_class,
            "escalation_risk_level": risk_level,
            "recommended_template_id": template_id,
            "tone_level": tone_level,
            "recommended_next_action": recommended_next_action,
            "agent_warning": agent_warning,
            "deescalation_phrases": deescalation_phrases,
            "forbidden_claims": forbidden_claims,
            "recommended_customer_reply": recommended_customer_reply,
            "template_selection_reason": template_selection_reason,
            "decision_reason": decision_reason,
        })

    out = pd.DataFrame(output_rows)

    # technische Sortierspalten
    out["_priority_sort"] = out["case_priority_class"].apply(priority_rank)
    out["_risk_sort"] = out["escalation_risk_level"].apply(risk_rank)
    out = out.sort_values(
        by=["_priority_sort", "_risk_sort", "priority_score"],
        ascending=[True, True, False]
    ).drop(columns=["_priority_sort", "_risk_sort"])

    agent_cols = [
        "case_id",
        "case_priority_class",
        "escalation_risk_level",
        "priority_score",
        "urgency_level",
        "business_value_level",
        "customer_name",
        "vehicle_model",
        "recommended_template_id",
        "tone_level",
        "recommended_next_action",
        "agent_warning",
        "deescalation_phrases",
        "forbidden_claims",
        "recommended_customer_reply",
        "template_selection_reason",
        "decision_reason",
    ]

    agent_view = out[[c for c in agent_cols if c in out.columns]].copy()

    run_info = pd.DataFrame([
        ["run_timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["input_file", str(input_file)],
        ["rulebook_file", str(rulebook_file)],
        ["processed_cases", len(out)],
        ["output_file", str(OUTPUT_FILE)],
    ], columns=["key", "value"])

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        out.to_excel(writer, sheet_name="Output_Simulation_MVP", index=False)
        agent_view.to_excel(writer, sheet_name="Agent_View_MVP", index=False)
        run_info.to_excel(writer, sheet_name="Run_Info", index=False)

    format_output_workbook(OUTPUT_FILE)

    log("")
    log("Engine erfolgreich ausgefÃ¼hrt.")
    log(f"Verarbeitete Cases: {len(out)}")
    log(f"Output erstellt: {OUTPUT_FILE}")
    log("")
    log("NÃ¤chster Schritt: Ã–ffne die Datei im Ordner 03_Output.")


if __name__ == "__main__":
    try:
        run_engine()
    except Exception as e:
        log("")
        log("FEHLER:")
        log(str(e))
        log("")
        log("PrÃ¼fe:")
        log("- Liegt eine Input-Datei in 01_Input?")
        log("- Liegt das Rulebook in 02_Regelmatrix?")
        log("- Stimmen die Spaltennamen der Input-Datei?")
        log("- Ist die Output-Datei gerade in Excel geÃ¶ffnet? Dann bitte schlieÃŸen.")
        sys.exit(1)



